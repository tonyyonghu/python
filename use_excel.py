from openpyxl import Workbook
from datetime import datetime
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors, Alignment

class ExcelUtil(object):
	"""
	pip install openpyxl
	pip install pillow
	"""
	def __init__(self):
		self.wb=Workbook()
		self.ws=self.wb.active
		#设置标题
		self.ws.title="您的表单"
		self.ws_two=self.wb.create_sheet("我的表单")
		self.ws_two.sheet_properties.tabColor="ff0000"
		self.ws_three=self.wb.create_sheet()
	def do_sth(self):
		self.ws['A1']=54
		self.ws['A2'] = "你好"
		self.ws['A3'] = datetime.now()

		for row in self.ws_two["A1:E5"]:
			for cell in row:
				cell.value=10
			#对数据进行求和
			self.ws_two['F1']='=SUM(A1:E1)'
		#设置文字
		font=Font(sz=20,color=colors.RED)
		self.ws['A1'].font=font
		#插入图片
		img=Image("./static/temp.gif")
		#openpyxl 向指定单元格添加图片并修改图片大小 以及修改单元格行高列宽
		# column_width=12.25
		# row_height=80.10
		# self.ws.column_dimensions['B'].width=column_width #修改列B的列宽
		# self.ws.row_dimensions[1].height = row_height     #修改行1的行高
		# newsize=(90,90)
		# img.width,img.height = newsize
		# self.ws.add_image(img,'B1')
		self.ws.merge_cells("A4:E5")#合并
		#取消合并
		# self.ws.unmerge_cells("A4:E5")
		self.wb.save("./static/test.xlsx")

	def read_xls(self):
		"""
			读取excel数据

		"""
		ws=load_workbook("./static/template.xlsx")
		names=ws.get_sheet_names()
		print(names)
		wb=ws.active

if __name__=="__main__":
	client=ExcelUtil()
	client.do_sth()