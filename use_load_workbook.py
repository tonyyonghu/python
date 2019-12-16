from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
import random

#北京大学高考数据统计   （文字居中红色）
#年份     最高分   平均分  （文字加粗）
#2008-2019 
#北京大学统计    


class ExcelUtils(object):
	def __init__(self):
		self.wb=Workbook()
		self.ws=self.wb.active
		self.ws.title="北京大学统计"

	def do_sth(self):
		self.ws.merge_cells("A1:C1")
		self.ws.cell(1,1).value="北京大学高考数据统计"
		font=Font(sz=16,color=colors.BLUE,name=u'宋体',bold=True)
		font1=Font(sz=12,color=colors.BLACK,name=u'宋体',bold=True)
		align=Alignment(horizontal='center',vertical='center')
		self.ws['A1'].font=font
		self.ws['A1'].alignment=align

		self.ws["A2"].value= "年份"
		self.ws['A2'].font=font1
		self.ws['A2'].alignment=align
		self.ws['B2'].value = "最高分"
		self.ws['B2'].font=font1
		self.ws['B2'].alignment=align
		self.ws['C2'].value = '平均分'
		self.ws['C2'].font=font1
		self.ws['C2'].alignment=align

		for x in range(3,15):
			self.ws['A'+str(x)].value = 2005+x
			self.ws['A'+str(x)].font=font1
			self.ws['A'+str(x)].alignment=align
			self.ws['B'+str(x)].value = random.randint(690,750)
			self.ws['B'+str(x)].font=font1
			self.ws['B'+str(x)].alignment=align
			self.ws['C'+str(x)].value = random.randint(659,695)
			self.ws['C'+str(x)].font=font1
			self.ws['C'+str(x)].alignment=align
		self.wb.save("./static/template.xlsx")

	def read_excel(self):
		ws=load_workbook("./static/template.xlsx")
		names=ws.get_sheet_names()
		print(names)
		wb=ws.active
		wb=ws[names[0]]
		for row in wb.rows:
			for cell in row:
				print(cell.value)


if __name__=="__main__":
	excel=ExcelUtils()
	excel.read_excel()
	# excel.do_sth()
	