from openpyxl import Workbook
import MySQLdb
from openpyxl import load_workbook

from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)

class GaokaoExport(object):
	def __init__(self):
		self.wb=load_workbook("./static/temp.xlsx")
		self.ws=self.wb.active
		self.ws.title="高考统计"
		self.ws.sheet_properties.tabColor="ff0000"

	def get_conn(self):
		conn=MySQLdb.connect(
			db="test",
			host="localhost",
			user='root',
			password="root",
			charset="utf8"
		)
		return conn

	def export_data(self):
		conn=self.get_conn()
		cursor=conn.cursor()
		#准备查询语句（如果数据量大，需要借助于分页查询）
		cursor.execute("select `year`,`max`,`avg` from user_score")
		rows=cursor.fetchall()

		row_id=10
		for i,row in enumerate(rows):
			(self.ws["C{0}".format(row_id)],self.ws["D{0}".format(row_id)],self.ws['E{0}'.format(row_id)])=row
			row_id+=1

		chart = AreaChart()
		chart.title = "统计表"
		chart.style = 13
		chart.x_axis.title = '年份'
		chart.y_axis.title = '分数'
		#横坐标
		cats = Reference(self.ws, min_col=3, min_row=10, max_row=row_id)
		#数据区域
		data = Reference(self.ws, min_col=4, min_row=9, max_col=5, max_row=row_id)
		chart.add_data(data, titles_from_data=True)
		chart.set_categories(cats)

		self.ws.add_chart(chart, "A{0}".format(row_id+5))

		# wb.save("area.xlsx")
		self.wb.save("./static/export_data.xlsx")

if __name__=="__main__":
	client=GaokaoExport()
	client.export_data()