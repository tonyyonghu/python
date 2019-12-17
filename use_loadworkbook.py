from openpyxl import load_workbook
from openpyxl import Workbook
import MySQLdb
class ExcelUtils(object):
	def __init__(self):
		self.wb=load_workbook("./static/template.xlsx")
		self.ws=self.wb.active

	def do_sth(self):
		all_row=self.ws.max_row   #最大行
		all_col=self.ws.max_column  #最大列
		conn=self.get_conn()
		cursor=conn.cursor()
		# sql='insert into `user_score`(`year`,`max`,`avg`) values(2001,400,460)'
		# cursor.execute(sql)
		# conn.autocommit(True)
		for x in range(3,all_row+1):
			year=self.ws['A'+str(x)].value
			max_score=self.ws['B'+str(x)].value
			avg_score=self.ws['C'+str(x)].value
			sql='insert into `user_score`(`year`,`max`,`avg`) values({year},{max_score},{avg_score})'.format(year=year,max_score=max_score,avg_score=avg_score)
			cursor.execute(sql)
			conn.autocommit(True)

	def get_conn(self):
		conn=MySQLdb.connect(
			db="test",
			host="localhost",
			user='root',
			password="root",
			charset="utf8"
		)
		return conn

	def export_xlsx(self):
		conn=self.get_conn()
		cursor=conn.cursor()
		#准备查询语句（如果数据量大，需要借助于分页查询）
		cursor.execute("select `year`,`max`,`avg` from user_score")
		rows=cursor.fetchall()

		wb=Workbook()
		ws=wb.active
		ws.title="北京大学高考统计"
		ws.merge_cells("A1:C1")
		ws['A1'] = "北京大学高考数据统计"
		ws['A2'] = "年份"
		ws['B2'] = "最高分"
		ws['C2'] = "平均分"
		for i,row in enumerate(rows):
			(ws["A"+str(i+3)],ws["B"+str(i+3)],ws['C'+str(i+3)])=row

		wb.save("./static/export.xlsx")


if __name__=="__main__":
	excel=ExcelUtils()
	#excel.do_sth()
	excel.export_xlsx()